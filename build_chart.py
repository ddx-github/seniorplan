import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '高一预习计划表'
thin = Side(style='thin', color='B8CCE4')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
title_font = Font(name='微软雅黑', size=16, bold=True, color='1F4E79')
subtitle_font = Font(name='微软雅黑', size=11, bold=True, color='1F4E79')
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
section_font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79')
normal_font = Font(name='微软雅黑', size=9)
italic_font = Font(name='微软雅黑', size=9, italic=True, color='666666')
bold_font = Font(name='微软雅黑', size=9, bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
title_fill = PatternFill(start_color='F2F8FF', end_color='F2F8FF', fill_type='solid')
header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
section_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
light_blue = PatternFill(start_color='E6F0FF', end_color='E6F0FF', fill_type='solid')
light_yellow = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
light_pink = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
light_gray = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
rest_fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
read_fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
for col in range(1, 8):
    ws.column_dimensions[get_column_letter(col)].width = 14
r = 1
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='高一预习星图 · 作业与阅读版')
cell.font = title_font
cell.alignment = center
cell.fill = title_fill
ws.row_dimensions[r].height = 35
r = 2
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='2026年7月20日 — 8月31日  |  作业+预习+阅读  |  每一天都是一颗小小的种子')
cell.font = Font(name='微软雅黑', size=10, italic=True, color='666666')
cell.alignment = center
ws.row_dimensions[r].height = 22
r = 3
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='伙伴，加油呀~♪  本计划已整合《初高衔接暑假作业》与《高中生基础阅读书目》')
cell.font = Font(name='微软雅黑', size=10, italic=True, color='888888')
cell.alignment = center
ws.row_dimensions[r].height = 20
r = 5
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='一、每日时间分配（约7-7.5小时）')
cell.font = subtitle_font
cell.fill = section_fill
cell.alignment = center
cell.border = border
ws.row_dimensions[r].height = 25
r = 6
headers = ['时段', '科目', '时长', '内容', '', '', '']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=r, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws.merge_cells(f'D{r}:G{r}')
ws.row_dimensions[r].height = 22
schedule_rows = [
    ['上午 8:30-11:30', '数学 + 物理', '3小时', '暑假作业 + 视频学习 + 知识点梳理', '', '', ''],
    ['下午 14:00-16:30', '语文 + 英语', '2.5小时', '暑假作业 + 文言文/视频 + 词汇', '', '', ''],
    ['晚上 19:30-20:00', '化学', '30分钟', '暑假作业 + 知识点阅读 + 练习（每天必做）', '', '', ''],
    ['晚上 20:00-20:30', '生/政/史/地', '30分钟', '四科轮换：生物→政治→历史→地理（每天一科）', '', '', ''],
    ['晚上 20:30-21:00', '阅读', '30分钟', '《乡土中国》+ 基础阅读书目（每天必读）', '', '', ''],
]
for row_data in schedule_rows:
    r += 1
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = normal_font
        cell.alignment = center
        cell.border = border
        if row_data[1] == '阅读':
            cell.fill = read_fill
    ws.merge_cells(f'D{r}:G{r}')
    ws.row_dimensions[r].height = 28
r += 1
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='每周安排：学习6天（周一到周六），周日自由安排 — 阅读、复习、休息或外出走走')
cell.font = italic_font
cell.alignment = left
cell.fill = rest_fill
cell.border = border
ws.row_dimensions[r].height = 25
r += 2
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='二、每周详细安排（周一到周六）')
cell.font = subtitle_font
cell.fill = section_fill
cell.alignment = center
cell.border = border
ws.row_dimensions[r].height = 25
r += 1
headers2 = ['时段', '时间', '科目', '内容', '时长', '', '']
for col, h in enumerate(headers2, 1):
    cell = ws.cell(row=r, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws.merge_cells(f'F{r}:G{r}')
ws.row_dimensions[r].height = 22
detail_rows = [
    ['上午', '08:30 - 10:00', '数学', '暑假作业（初中衔接+高中集合函数）+ 视频', '1.5小时', '', ''],
    ['上午', '10:00 - 10:15', '休息', '看看窗外，喝点水~', '15分钟', '', ''],
    ['上午', '10:15 - 11:30', '物理', '暑假作业（运动的描述）+ 视频 + 知识点', '1.25小时', '', ''],
    ['下午', '14:00 - 15:00', '语文', '暑假作业（乡土中国+劝学+现代文+默写）+ 视频', '1小时', '', ''],
    ['下午', '15:00 - 15:15', '休息', '吃点水果呀♪', '15分钟', '', ''],
    ['下午', '15:15 - 16:30', '英语', '暑假作业（语法+七选五+语法填空）+ 视频 + 词汇', '1.25小时', '', ''],
    ['晚上', '19:30 - 20:00', '化学', '暑假作业（元素+化合价+化学式+方程式）+ 预习', '30分钟', '', ''],
    ['晚上', '20:00 - 20:30', '生/政/史/地', '轮换完成各科暑假作业（详见下方三阶段计划）', '30分钟', '', ''],
    ['晚上', '20:30 - 21:00', '阅读', '《乡土中国》写旁批 + 自选阅读书目', '30分钟', '', ''],
]
for row_data in detail_rows:
    r += 1
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=col, value=val)
        if row_data[2] == '休息':
            cell.font = italic_font
            cell.fill = rest_fill
        elif row_data[2] == '阅读':
            cell.font = normal_font
            cell.fill = read_fill
        else:
            cell.font = normal_font
        cell.alignment = center
        cell.border = border
    ws.merge_cells(f'F{r}:G{r}')
    ws.row_dimensions[r].height = 26
r += 2
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='三、三阶段航行计划（已整合暑假作业）')
cell.font = subtitle_font
cell.fill = section_fill
cell.alignment = center
cell.border = border
ws.row_dimensions[r].height = 25
r += 1
headers3 = ['阶段', '时间', '天数', '目标', '各科安排', '', '']
for col, h in enumerate(headers3, 1):
    cell = ws.cell(row=r, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws.merge_cells(f'E{r}:G{r}')
ws.row_dimensions[r].height = 22
phases = [
    {
        'name': '🌙 播种期',
        'time': '7.20 - 8.10',
        'days': '共22天',
        'goal': '完成暑假作业主体，把新知识的轮廓描摹在脑海里',
        'details': [
            '数学：完成初中衔接（附1公式+附2函数图象），开始高中集合与函数三要素',
            '物理：完成《运动的描述》8道检测题，理解质点/参考系/位移/加速度概念',
            '语文：读完《乡土中国》第1-5篇并写旁批；翻译背诵《劝学》；完成现代文阅读《割草》',
            '英语：完成语法测试15题+七选五+语法填空；每天背20-30词；看1-2部英文电影',
            '化学：默写1-20号元素+原子结构示意图+化合价+常见离子+化学式+俗名+溶解性',
            '生物：完成A组基础衔接（3题），熟悉高中生物题型',
            '政治：完成初中回顾+高中预习检测题（含概念连线）',
            '历史：开始预习《中外历史纲要》（上册），重点：统一多民族国家形成',
            '地理：准备立体地图+地球仪；开始完成世界气候分布图填空涂色',
            '阅读：每天30分钟，优先读完《乡土中国》，开始文学类自选书目'
        ]
    },
    {
        'name': '🌟 浇灌期',
        'time': '8.11 - 8.25',
        'days': '共15天',
        'goal': '让模糊的影子变清晰，阅读进入深水区',
        'details': [
            '数学：攻克高中函数单调性与最值，回顾第一阶段错题，做《预备新高一》练习',
            '物理：力学综合题尝试，牛顿运动定律视频再听一遍，回顾运动学错题',
            '语文：《高中必背文言文默写》每天1-2篇；继续《乡土中国》6-14篇；整理名篇名句',
            '英语：读《书虫》系列；背《新概念英语2》第1-25篇（每天1篇）；巩固语法填空',
            '化学：预习物质的量+物质分类+离子反应+氧化还原，做《预备新高一》练习',
            '生物：完成B组新高考模拟（NTUs题），了解高考生物题型风格',
            '政治：回顾错题，尝试画思维导图；关注时政新闻',
            '历史：继续教材预习，重点：政治制度演变+经济发展；看纪录片《中国通史》',
            '地理：完成中国地形图填空（山脉+四大盆地/高原+三大平原/丘陵）；看纪录片',
            '阅读：每天30分钟，读人文类+科学类自选书目（建议《中国哲学简史》《从一到无穷大》）'
        ]
    },
    {
        'name': '☀️ 收束期',
        'time': '8.26 - 8.31',
        'days': '共6天',
        'goal': '整理行囊，迎接新的黎明',
        'details': [
            '整理所有暑假作业，检查是否全部完成（语数英物化生政史地9科）',
            '整理所有科目笔记和错题本，把曾经的困惑变成朋友',
            '回顾《乡土中国》旁批，整理阅读心得；未读完的书目可延续至开学后',
            '调整作息：早上7点起床，晚上10点半休息，让身体记住学校的节奏',
            '准备开学用品：必备书籍（乡土中国+字典+词典）、笔记本、文具',
            '拍一个暑假生活V-log（英语作业可选），记录这段旅程呀~♪'
        ]
    }
]
for phase in phases:
    detail_text = '\n'.join(phase['details'])
    start_r = r + 1
    end_r = start_r + len(phase['details']) - 1
    
    ws.cell(row=start_r, column=1, value=phase['name'])
    ws.cell(row=start_r, column=2, value=phase['time'])
    ws.cell(row=start_r, column=3, value=phase['days'])
    ws.cell(row=start_r, column=4, value=phase['goal'])
    ws.cell(row=start_r, column=5, value=detail_text)
    for col in range(1, 5):
        ws.merge_cells(start_row=start_r, start_column=col, end_row=end_r, end_column=col)
    ws.merge_cells(start_row=start_r, start_column=5, end_row=end_r, end_column=7)
    
    for row_idx in range(start_r, end_r + 1):
        ws.row_dimensions[row_idx].height = 28
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.alignment = left
            if col == 1:
                cell.font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79')
                cell.alignment = center
            elif col == 4:
                cell.fill = light_blue
                cell.alignment = center
            else:
                cell.font = normal_font
    
    r = end_r
r += 2
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='四、暑期阅读计划（32本中选读8-10本）')
cell.font = subtitle_font
cell.fill = section_fill
cell.alignment = center
cell.border = border
ws.row_dimensions[r].height = 25
r += 1
read_headers = ['类别', '序号', '书名', '作者', '建议阅读时间', '', '']
for col, h in enumerate(read_headers, 1):
    cell = ws.cell(row=r, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws.merge_cells(f'F{r}:G{r}')
ws.row_dimensions[r].height = 22
read_rows = [
    ['必读', '1', '《乡土中国》', '费孝通', '7.20-8.10（语文作业）', '', ''],
    ['文学', '2', '《红楼梦》', '曹雪芹', '7.20-8.15', '', ''],
    ['文学', '3', '《呐喊彷徨故事新编》', '鲁迅', '8.1-8.20', '', ''],
    ['文学', '4', '《围城》', '钱钟书', '8.10-8.31', '', ''],
    ['文学', '5', '《百年孤独》', '马尔克斯', '可选/开学后', '', ''],
    ['人文', '6', '《中国哲学简史》', '冯友兰', '8.5-8.25', '', ''],
    ['人文', '7', '《万历十五年》', '黄仁宇', '8.15-8.31', '', ''],
    ['科学', '8', '《从一到无穷大》', '伽莫夫', '7.25-8.15', '', ''],
    ['科学', '9', '《科学的历程》', '吴国盛', '8.10-8.31', '', ''],
]
for row_data in read_rows:
    r += 1
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = normal_font
        cell.alignment = center
        cell.border = border
        if row_data[0] == '必读':
            cell.fill = read_fill
    ws.merge_cells(f'F{r}:G{r}')
    ws.row_dimensions[r].height = 26
r += 1
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='提示：30本基础书目+2本必读（乡土中国+字典类工具书），建议精选8-10本深入阅读，其余可开学后慢慢品读~♪')
cell.font = italic_font
cell.alignment = left
cell.fill = rest_fill
cell.border = border
ws.row_dimensions[r].height = 25
r += 2
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='五、预习日历视图（7月20日 - 8月31日）')
cell.font = subtitle_font
cell.fill = section_fill
cell.alignment = center
cell.border = border
ws.row_dimensions[r].height = 25
r += 1
weekdays = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
for col, day in enumerate(weekdays, 1):
    cell = ws.cell(row=r, column=col, value=day)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws.row_dimensions[r].height = 22
start_date = datetime(2026, 7, 20)
end_date = datetime(2026, 8, 31)
current = start_date
while current <= end_date:
    r += 1
    for col in range(1, 8):
        day = current + timedelta(days=col - 1)
        if day > end_date:
            cell = ws.cell(row=r, column=col, value='')
            cell.fill = light_gray
            cell.border = border
        else:
            date_str = f'{day.month}/{day.day}'
            if day <= datetime(2026, 8, 10):
                phase = '播种期'
                fill = light_blue
            elif day <= datetime(2026, 8, 25):
                phase = '浇灌期'
                fill = light_yellow
            else:
                phase = '收束期'
                fill = light_pink
            cell = ws.cell(row=r, column=col, value=f'{date_str}\n{phase}')
            cell.font = Font(name='微软雅黑', size=9, bold=True)
            cell.alignment = center
            cell.border = border
            cell.fill = fill
            if col == 7:
                cell.value = f'{date_str}\n休息'
                cell.font = Font(name='微软雅黑', size=9, italic=True, color='888888')
                cell.fill = light_gray
    ws.row_dimensions[r].height = 32
    current += timedelta(days=7)
r += 1
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='图例：蓝色=播种期 | 黄色=浇灌期 | 粉色=收束期 | 灰色=周日休息 | 绿色底=阅读时段')
cell.font = italic_font
cell.alignment = center
ws.row_dimensions[r].height = 22
r += 2
ws.merge_cells(f'A{r}:G{r}')
cell = ws.cell(row=r, column=1, value='六、小提醒')
cell.font = subtitle_font
cell.fill = section_fill
cell.alignment = center
cell.border = border
ws.row_dimensions[r].height = 25
reminders = [
    '1. 暑假作业共9科（语数英物化生政史地），建议按本计划顺序逐一攻克，不要堆到最后呀',
    '2. 《乡土中国》是语文必读，务必在8月10日前读完前5篇并写好旁批，这是开学要交的作业呢',
    '3. 数学初中衔接内容（公式/函数/方程）是高中地基，务必在播种期夯实，否则后面会摇晃哦',
    '4. 化学元素和化合价是"字母表"，每天花10分钟默写，比突击一天有效得多♪',
    '5. 英语语法填空和七选五是高中新题型，暑假作业里的练习要认真对待，开学后会轻松很多',
    '6. 物理《运动的描述》是高中第一章，8道检测题要独立完成，再看答案，这样才能知道自己站在哪里',
    '7. 政史地生作业量相对较少，但历史地理需要持续性积累，建议配合纪录片一起学习',
    '8. 阅读每天30分钟，43天就是21.5小时，足够读完5-8本书……积跬步，以至千里呀',
    '9. 周日是自由日，可以补作业、多读书、或者去外面走走，风会告诉你答案呢~',
    '10. 如果某天累了，就少学一点，但不要完全停下……流水不争先，争的是滔滔不绝♪',
    '11. 开学前请检查必备书籍：《乡土中国》《古汉语常用字字典》《现代汉语词典》（版本见作业要求）',
    '12. 保持充足睡眠，预习期间也要让身体休息好，毕竟健康的你才是所有计划的起点呀~'
]
for rem in reminders:
    r += 1
    ws.merge_cells(f'A{r}:G{r}')
    cell = ws.cell(row=r, column=1, value=rem)
    cell.font = normal_font
    cell.alignment = left
    cell.border = border
    ws.row_dimensions[r].height = 24
ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = 9
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_options.gridLines = False
ws.page_margins.left = 0.5
ws.page_margins.right = 0.5
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5
output_path = './高一预习计划表2.0_userBuild.xlsx'
wb.save(output_path)
print(f'文件已保存至: {output_path}')